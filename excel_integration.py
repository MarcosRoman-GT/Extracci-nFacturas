import os
import openpyxl
from openpyxl.utils import get_column_letter

# Ruta al archivo Excel existente
EXCEL_PATH = r"C:\Users\vqzjo\OneDrive\Escritorio\Proyecto con extracción de logo\Extracci-nFacturas\DB.xlsx"

# Función para insertar una fila de datos en el Excel
def append_to_excel(data: dict):
    """
    Agrega los datos de una factura al final de la hoja activa de un archivo Excel existente.
    """
    # Si no existe el archivo, lanzar excepción o crear encabezados manualmente
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"No se encontró el archivo Excel: {EXCEL_PATH}")

    # Carga el libro y la hoja activa
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Lee encabezados de la primera fila
    headers = [cell.value for cell in ws[1]]
    # Construye la fila de valores según el orden de los encabezados
    row = [data.get(str(header), '') for header in headers]

    # Agrega la fila al final
    ws.append(row)

    # Ajuste de ancho de columnas (opcional)
    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(15, len(str(header)) + 2)

    # Guarda los cambios
    wb.save(EXCEL_PATH)
